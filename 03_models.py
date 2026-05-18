"""
03_models.py
────────────
【這個檔案在做什麼】
吃進 features.csv，對 6 個模型做 LOOCV (Leave-One-Out Cross-Validation)，
輸出：
  results/benchmark_table.xlsx  — 模型比較表（RMSE + R² + bootstrap 95% CI）
  figures/benchmark_bar.png     — RMSE 比較長條圖（履歷用）

【為什麼用 LOOCV】
樣本只有 ~120-170 隻眼（移除缺失後更少）。
LOOCV 是小樣本標準做法：每次用 n-1 隻眼訓練，對剩下 1 隻眼預測，
重複 n 次後算整體 RMSE/R²。這樣不浪費任何一個樣本。

【6 個模型選擇邏輯】
Linear        — 基準線（最簡單，解釋性最高）
Ridge / Lasso — 加入正則化（處理特徵間共線性，Lasso 同時做特徵選擇）
Random Forest — 非線性集成模型（不假設線性關係）
XGBoost       — Gradient Boosting（業界主流，通常 RMSE 最低）
SVR           — Support Vector Regression（對小樣本有理論優勢）

【特徵策略】
early_slope_3m 缺失率 51%，用兩組特徵集：
  feat_set_A：去掉 early_slope_3m，用 early_slope_6m（更多樣本）
  feat_set_B：只用有 early_slope_3m 的子集（樣本更少但特徵更完整）
主要報告 feat_set_A 結果。
"""

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from pathlib import Path

from sklearn.linear_model    import LinearRegression, Ridge, Lasso
from sklearn.ensemble        import RandomForestRegressor
from sklearn.svm             import SVR
from sklearn.preprocessing   import StandardScaler
from sklearn.pipeline        import Pipeline
from sklearn.model_selection import LeaveOneOut
from sklearn.metrics         import mean_squared_error, r2_score
from sklearn.impute          import SimpleImputer

try:
    from xgboost import XGBRegressor
    HAS_XGB = True
except ImportError:
    HAS_XGB = False
    print("⚠️  xgboost 未安裝，跳過 XGBoost")

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE      = Path(__file__).parent
FEAT_FILE = BASE / "features.csv"
RES_DIR   = BASE / "results"
FIG_DIR   = BASE / "figures"
RES_DIR.mkdir(exist_ok=True)
FIG_DIR.mkdir(exist_ok=True)


# ── 特徵欄位 ──────────────────────────────────────────────────────
FEATURES_A = [
    "age_start", "sex", "laterality",
    "myopia_d", "axl_baseline", "axl_dev_from_norm",
    "mi_rx", "atropine_conc",
    "early_slope_6m", "axl_std_6m",
]
TARGET = "y_annual_mm_yr"


# ── 模型定義 ──────────────────────────────────────────────────────
def get_models():
    models = {
        "Linear Regression": Pipeline([
            ("imp", SimpleImputer(strategy="median")),
            ("scl", StandardScaler()),
            ("mdl", LinearRegression()),
        ]),
        "Ridge": Pipeline([
            ("imp", SimpleImputer(strategy="median")),
            ("scl", StandardScaler()),
            ("mdl", Ridge(alpha=1.0)),
        ]),
        "Lasso": Pipeline([
            ("imp", SimpleImputer(strategy="median")),
            ("scl", StandardScaler()),
            ("mdl", Lasso(alpha=0.01, max_iter=5000)),
        ]),
        "Random Forest": Pipeline([
            ("imp", SimpleImputer(strategy="median")),
            ("mdl", RandomForestRegressor(
                n_estimators=200, max_depth=4,
                random_state=42, n_jobs=-1
            )),
        ]),
        "SVR": Pipeline([
            ("imp", SimpleImputer(strategy="median")),
            ("scl", StandardScaler()),
            ("mdl", SVR(kernel="rbf", C=1.0, epsilon=0.05)),
        ]),
    }
    if HAS_XGB:
        models["XGBoost"] = Pipeline([
            ("imp", SimpleImputer(strategy="median")),
            ("mdl", XGBRegressor(
                n_estimators=200, max_depth=3,
                learning_rate=0.05, subsample=0.8,
                random_state=42, verbosity=0,
            )),
        ])
    return models


# ── LOOCV ─────────────────────────────────────────────────────────
def run_loocv(X: np.ndarray, y: np.ndarray, models: dict) -> pd.DataFrame:
    loo = LeaveOneOut()
    results = []

    for name, model in models.items():
        preds = np.zeros(len(y))
        for train_idx, test_idx in loo.split(X):
            X_tr, X_te = X[train_idx], X[test_idx]
            y_tr        = y[train_idx]
            model.fit(X_tr, y_tr)
            preds[test_idx] = model.predict(X_te)

        rmse = np.sqrt(mean_squared_error(y, preds))
        r2   = r2_score(y, preds)

        # Bootstrap 95% CI on RMSE
        np.random.seed(42)
        boot_rmse = []
        for _ in range(1000):
            idx = np.random.choice(len(y), len(y), replace=True)
            boot_rmse.append(np.sqrt(mean_squared_error(y[idx], preds[idx])))
        ci_lo, ci_hi = np.percentile(boot_rmse, [2.5, 97.5])

        results.append({
            "Model"       : name,
            "RMSE (mm/yr)": round(rmse, 4),
            "R²"          : round(r2, 4),
            "CI_lo"       : round(ci_lo, 4),
            "CI_hi"       : round(ci_hi, 4),
            "95% CI"      : f"[{ci_lo:.3f}, {ci_hi:.3f}]",
        })
        print(f"  {name:<22} RMSE={rmse:.4f}  R²={r2:.4f}  CI=[{ci_lo:.3f},{ci_hi:.3f}]")

    return pd.DataFrame(results).sort_values("RMSE (mm/yr)")


# ── Excel 輸出（漂亮格式） ─────────────────────────────────────────
def save_excel(df: pd.DataFrame, n_samples: int, path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Model Benchmark"

    # 標題
    ws["A1"] = "Myopia ML — Model Benchmark (LOOCV)"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Feature set A  |  n = {n_samples} eyes  |  Target: annualized AXL progression (mm/yr)"
    ws["A2"].font = Font(italic=True, size=10, color="595959")
    ws.merge_cells("A2:F2")

    # 表頭
    headers = ["Rank", "Model", "RMSE (mm/yr)", "R²", "Bootstrap 95% CI", "Note"]
    header_fill  = PatternFill("solid", start_color="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF")
    thin_border  = Border(
        bottom=Side(style="thin", color="AAAAAA"),
        right =Side(style="thin", color="AAAAAA"),
    )

    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_i, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 資料列
    best_rmse = df["RMSE (mm/yr)"].min()
    row_fills = {
        0: "D6E4F0",   # 最佳 — 淡藍
        1: "EBF5D0",   # 第二 — 淡綠
    }

    for rank_i, (_, row) in enumerate(df.iterrows()):
        excel_row = rank_i + 5
        note = "★ Best" if row["RMSE (mm/yr)"] == best_rmse else ""
        values = [
            rank_i + 1,
            row["Model"],
            row["RMSE (mm/yr)"],
            row["R²"],
            row["95% CI"],
            note,
        ]
        fill_hex = row_fills.get(rank_i, "FFFFFF")
        fill_obj = PatternFill("solid", start_color=fill_hex)

        for col_i, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_i, value=val)
            cell.fill      = fill_obj
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal="center" if col_i != 2 else "left")
            if col_i == 3:
                cell.number_format = "0.0000"
            if col_i == 4:
                cell.number_format = "0.0000"

    # 欄寬
    for col, width in zip("ABCDEF", [7, 22, 14, 10, 22, 10]):
        ws.column_dimensions[col].width = width

    # 備注
    note_row = len(df) + 7
    ws.cell(row=note_row, column=1,
            value="Note: LOOCV = Leave-One-Out Cross-Validation. "
                  "Bootstrap CI computed from 1000 resamples.")
    ws.cell(row=note_row, column=1).font = Font(italic=True, size=9, color="888888")
    ws.merge_cells(f"A{note_row}:F{note_row}")

    wb.save(path)
    print(f"✓ Excel 輸出：{path}")


# ── 視覺化 ────────────────────────────────────────────────────────
def save_figure(df: pd.DataFrame, path: Path):
    fig, ax = plt.subplots(figsize=(8, 4.5))
    colors = ["#1F4E79", "#2E75B6", "#5BA3D9", "#9DC3E6", "#C5DCF0", "#E2EFF8"]
    colors = colors[:len(df)]

    bars = ax.barh(df["Model"], df["RMSE (mm/yr)"], color=colors, edgecolor="white", height=0.55)

    # 誤差棒
    xerr_lo = df["RMSE (mm/yr)"] - df["CI_lo"]
    xerr_hi = df["CI_hi"] - df["RMSE (mm/yr)"]
    ax.errorbar(
        df["RMSE (mm/yr)"], df["Model"],
        xerr=[xerr_lo, xerr_hi],
        fmt="none", color="#333333", capsize=4, linewidth=1.2
    )

    # 數值標籤
    for bar, rmse in zip(bars, df["RMSE (mm/yr)"]):
        ax.text(
            rmse + 0.002, bar.get_y() + bar.get_height() / 2,
            f"{rmse:.4f}", va="center", ha="left", fontsize=9, color="#1F4E79"
        )

    ax.set_xlabel("RMSE (mm/yr)", fontsize=11)
    ax.set_title(
        "Model Benchmark — LOOCV RMSE\n"
        "Myopia Axial Length Progression Prediction",
        fontsize=12, fontweight="bold", pad=12
    )
    ax.invert_yaxis()
    ax.set_xlim(0, df["CI_hi"].max() * 1.18)
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(axis="x", linestyle="--", alpha=0.4)

    plt.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"✓ 圖表輸出：{path}")


# ── 主程式 ────────────────────────────────────────────────────────
def main():
    df = pd.read_csv(FEAT_FILE)
    print(f"讀入 features.csv：{len(df)} 隻眼")

    # 特徵集 A：移除 early_slope_3m（缺失率太高）
    feats = [f for f in FEATURES_A if f in df.columns]
    sub   = df[feats + [TARGET]].dropna(subset=[TARGET])
    print(f"移除 target 缺失後：{len(sub)} 隻眼")

    X = sub[feats].values
    y = sub[TARGET].values

    print(f"\n跑 LOOCV（n={len(y)}）…")
    models  = get_models()
    results = run_loocv(X, y, models)

    print("\n=== 結果 ===")
    print(results[["Model", "RMSE (mm/yr)", "R²", "95% CI"]].to_string(index=False))

    save_excel(results, n_samples=len(y), path=RES_DIR / "benchmark_table.xlsx")
    save_figure(results, path=FIG_DIR / "benchmark_bar.png")


if __name__ == "__main__":
    main()
